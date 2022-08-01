from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image

im = Image.open('image.png')
rgb_im = im.convert('RGB')
rgbArr = []

def rgbToHex(rgb):
    return '%02x%02x%02x' % rgb

def rgbArray():
    imgX = im.size[0]
    imgY = im.size[1]

    for i in range(imgY):
        rgbArr.append([])

    for i in range(imgX):
        for j in range(imgY):
            r, g, b = rgb_im.getpixel((i, j))
        
            rgbArr[j].append([r, g, b])

    return rgbArr

def generateSheet():
    wb = Workbook()

    ws =  wb.active
    ws.title = "image"

    for i in range(im.size[0]):
        ws.column_dimensions[get_column_letter(i + 1)].width = 3.0
        for j in range(im.size[1]):
            cell = get_column_letter(i + 1) + str(j + 1)
            ws[cell].fill = PatternFill(fgColor = rgbArr[j][i], fill_type = "solid")

    wb.save(filename = 'image.xlsx')

rgbArray()

for i in range(len(rgbArr)):
    for j in range(len(rgbArr[i])):
        rgbArr[i][j] = rgbToHex((rgbArr[i][j][0], rgbArr[i][j][1], rgbArr[i][j][2]))

print('this tool currently only works with .png files, large files may also take long to process/not work entirely')
generateSheet()

print('\ndone')
